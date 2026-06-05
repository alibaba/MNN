import builtins
import contextlib
import io
import json
import os
import py_compile
import sys
import time
import traceback
from os.path import abspath, commonpath, isabs, join, realpath


ALLOWED_IMPORT_ROOTS = {
    "base64",
    "bisect",
    "calendar",
    "collections",
    "csv",
    "dateutil",
    "datetime",
    "decimal",
    "et_xmlfile",
    "fractions",
    "functools",
    "hashlib",
    "heapq",
    "html",
    "io",
    "itertools",
    "json",
    "math",
    "numpy",
    "operator",
    "openpyxl",
    "pandas",
    "pytz",
    "random",
    "re",
    "statistics",
    "string",
    "textwrap",
    "time",
    "tzdata",
    "urllib",
    "xml",
}

DENIED_IMPORT_ROOTS = {
    "ctypes",
    "multiprocessing",
    "os",
    "pathlib",
    "shutil",
    "socket",
    "subprocess",
    "sys",
}


def run_code(code, input_text="", timeout_ms=15000, workspace_dir=""):
    start = time.monotonic()
    deadline = start + max(1, min(int(timeout_ms or 15000), 30000)) / 1000.0
    stdout = io.StringIO()
    stderr = io.StringIO()
    state = {"result": None, "has_result": False}

    def check_deadline():
        if time.monotonic() > deadline:
            raise TimeoutError("python_exec exceeded timeout")

    def trace_func(frame, event, arg):
        check_deadline()
        return trace_func

    original_import = builtins.__import__

    def safe_import(name, globals=None, locals=None, fromlist=(), level=0):
        root = name.split(".", 1)[0]
        caller_file = str((globals or {}).get("__file__", ""))
        is_package_internal = caller_file and not caller_file.startswith("<") and not caller_file.endswith("agent_python.py")
        if is_package_internal:
            return original_import(name, globals, locals, fromlist, level)
        if root in DENIED_IMPORT_ROOTS or root not in ALLOWED_IMPORT_ROOTS:
            raise ImportError("module is not allowed: " + name)
        return original_import(name, globals, locals, fromlist, level)

    def set_result(value):
        state["result"] = value
        state["has_result"] = True

    workspace_root = realpath(workspace_dir) if workspace_dir else ""
    files_before = _snapshot_workspace_files(workspace_root)

    def validate_workspace_path(path):
        if not path:
            raise PermissionError("empty path")
        raw_path = str(path)
        if workspace_root and not isabs(raw_path):
            raw_path = join(workspace_root, raw_path)
        full = realpath(abspath(raw_path))
        if not workspace_root or commonpath([workspace_root, full]) != workspace_root:
            raise PermissionError("path is outside python workspace: " + str(path))
        return full

    def safe_open(file, mode="r", *args, **kwargs):
        return open(validate_workspace_path(file), mode, *args, **kwargs)

    def read_excel(path, max_rows=200, max_sheets=10, values_only=True):
        full_path = validate_workspace_path(path)
        return _read_excel(full_path, max_rows=max_rows, max_sheets=max_sheets, values_only=values_only)

    def write_excel(filename, sheets):
        full_path = validate_workspace_path(filename)
        return _write_excel(full_path, sheets)

    def script_path(name):
        safe_name = _safe_script_name(name)
        script_dir = validate_workspace_path("python")
        os.makedirs(script_dir, exist_ok=True)
        return validate_workspace_path(join("python", safe_name))

    def save_script(name, source):
        full_path = script_path(name)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        with open(full_path, "w", encoding="utf-8") as f:
            f.write(str(source or ""))
        return {"path": full_path, "bytes": len(str(source or "").encode("utf-8"))}

    def load_script(name):
        full_path = script_path(name)
        with open(full_path, "r", encoding="utf-8") as f:
            return {"path": full_path, "source": f.read()}

    def list_scripts():
        script_dir = validate_workspace_path("python")
        os.makedirs(script_dir, exist_ok=True)
        scripts = []
        for root, _, files in os.walk(script_dir):
            for filename in files:
                if not filename.endswith(".py"):
                    continue
                full_path = join(root, filename)
                rel = os.path.relpath(full_path, script_dir).replace("\\", "/")
                scripts.append({"name": rel, "path": full_path, "bytes": os.path.getsize(full_path)})
        return sorted(scripts, key=lambda item: item["name"])

    def run_script(name):
        full_path = script_path(name)
        with open(full_path, "r", encoding="utf-8") as f:
            source = f.read()
        compiled_script = compile(source, full_path, "exec")
        exec(compiled_script, globals_dict, globals_dict)
        return state["result"] if state["has_result"] else globals_dict.get("result", None)

    def compile_script(name):
        full_path = script_path(name)
        try:
            py_compile.compile(full_path, doraise=True)
            return {"ok": True, "path": full_path, "error": ""}
        except py_compile.PyCompileError as exc:
            return {"ok": False, "path": full_path, "error": str(exc)}

    safe_builtins = {
        "abs": abs,
        "all": all,
        "any": any,
        "bool": bool,
        "bytes": bytes,
        "chr": chr,
        "dict": dict,
        "divmod": divmod,
        "enumerate": enumerate,
        "filter": filter,
        "float": float,
        "format": format,
        "hash": hash,
        "hex": hex,
        "int": int,
        "isinstance": isinstance,
        "issubclass": issubclass,
        "len": len,
        "list": list,
        "map": map,
        "max": max,
        "min": min,
        "next": next,
        "ord": ord,
        "pow": pow,
        "print": print,
        "range": range,
        "repr": repr,
        "reversed": reversed,
        "round": round,
        "set": set,
        "slice": slice,
        "sorted": sorted,
        "str": str,
        "sum": sum,
        "tuple": tuple,
        "zip": zip,
        "Exception": Exception,
        "ValueError": ValueError,
        "TypeError": TypeError,
        "RuntimeError": RuntimeError,
        "TimeoutError": TimeoutError,
        "__import__": safe_import,
        "open": safe_open,
    }

    globals_dict = {
        "__builtins__": safe_builtins,
        "__name__": "__agent_python__",
        "input_text": input_text or "",
        "workspace_dir": workspace_root,
        "read_excel": read_excel,
        "write_excel": write_excel,
        "save_script": save_script,
        "load_script": load_script,
        "list_scripts": list_scripts,
        "run_script": run_script,
        "compile_script": compile_script,
        "emit": set_result,
        "set_result": set_result,
    }

    try:
        globals_dict["input_json"] = json.loads(input_text) if input_text else None
    except Exception:
        globals_dict["input_json"] = None

    ok = True
    error = ""
    previous_cwd = os.getcwd()
    sys.settrace(trace_func)
    try:
        if workspace_root:
            os.makedirs(workspace_root, exist_ok=True)
            os.chdir(workspace_root)
        compiled = compile(code or "", "<agent_python>", "exec")
        with contextlib.redirect_stdout(stdout), contextlib.redirect_stderr(stderr):
            exec(compiled, globals_dict, globals_dict)
    except Exception:
        ok = False
        error = traceback.format_exc(limit=6)
    finally:
        sys.settrace(None)
        try:
            os.chdir(previous_cwd)
        except Exception:
            pass

    result = state["result"] if state["has_result"] else globals_dict.get("result", None)
    generated_files = _collect_generated_files(workspace_root, files_before)
    return json.dumps(
        {
            "ok": ok,
            "stdout": stdout.getvalue(),
            "stderr": stderr.getvalue(),
            "error": error,
            "result": _json_safe(result),
            "files": generated_files,
            "elapsed_ms": int((time.monotonic() - start) * 1000),
        },
        ensure_ascii=False,
    )


def _json_safe(value):
    try:
        json.dumps(value)
        return value
    except Exception:
        return repr(value)


def _snapshot_workspace_files(workspace_root):
    snapshot = {}
    if not workspace_root or not os.path.isdir(workspace_root):
        return snapshot
    for root, dirs, files in os.walk(workspace_root):
        dirs[:] = [d for d in dirs if d != "__pycache__"]
        for filename in files:
            full_path = realpath(join(root, filename))
            try:
                rel_path = os.path.relpath(full_path, workspace_root).replace("\\", "/")
                stat = os.stat(full_path)
                snapshot[rel_path] = (stat.st_mtime_ns, stat.st_size)
            except Exception:
                continue
    return snapshot


def _guess_mime_type(path):
    lower = path.lower()
    if lower.endswith(".xlsx"):
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if lower.endswith(".xls"):
        return "application/vnd.ms-excel"
    if lower.endswith(".csv"):
        return "text/csv"
    if lower.endswith(".pdf"):
        return "application/pdf"
    if lower.endswith(".txt"):
        return "text/plain"
    if lower.endswith(".json"):
        return "application/json"
    if lower.endswith(".png"):
        return "image/png"
    if lower.endswith(".jpg") or lower.endswith(".jpeg"):
        return "image/jpeg"
    return "application/octet-stream"


def _collect_generated_files(workspace_root, before):
    if not workspace_root or not os.path.isdir(workspace_root):
        return []
    after = _snapshot_workspace_files(workspace_root)
    changed = []
    for rel_path, stat in after.items():
        if rel_path.startswith("python/") or "/__pycache__/" in rel_path:
            continue
        if before.get(rel_path) == stat:
            continue
        full_path = realpath(join(workspace_root, rel_path))
        changed.append(
            {
                "name": os.path.basename(full_path),
                "path": full_path,
                "relative_path": rel_path,
                "mime_type": _guess_mime_type(full_path),
                "size_bytes": stat[1],
            }
        )
    changed.sort(key=lambda item: item["relative_path"])
    return changed[:20]


def _read_excel(path, max_rows=200, max_sheets=10, values_only=True):
    from datetime import date, datetime, time as dt_time
    from decimal import Decimal
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = []
        for sheet_name in workbook.sheetnames[: max(1, int(max_sheets or 10))]:
            ws = workbook[sheet_name]
            rows = []
            for row_index, row in enumerate(ws.iter_rows(values_only=values_only), start=1):
                if row_index > max(1, int(max_rows or 200)):
                    break
                rows.append([_cell_value(v) for v in row])
            sheets.append(
                {
                    "name": sheet_name,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "rows_returned": len(rows),
                    "rows": rows,
                }
            )
        return {"path": path, "sheet_count": len(workbook.sheetnames), "sheets": sheets}
    finally:
        workbook.close()


def _cell_value(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _safe_script_name(name):
    raw = str(name or "").strip().replace("\\", "/")
    if not raw:
        raise ValueError("script name is empty")
    parts = [part for part in raw.split("/") if part not in ("", ".", "..")]
    safe_parts = []
    for part in parts:
        cleaned = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in part)
        if cleaned:
            safe_parts.append(cleaned)
    if not safe_parts:
        raise ValueError("invalid script name")
    filename = "/".join(safe_parts)
    if not filename.endswith(".py"):
        filename += ".py"
    return filename


def _write_excel(path, sheets):
    from openpyxl import Workbook

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    if isinstance(sheets, dict):
        iterable = sheets.items()
    elif isinstance(sheets, list):
        iterable = []
        for index, item in enumerate(sheets, start=1):
            if isinstance(item, dict):
                iterable.append((item.get("name") or ("Sheet" + str(index)), item.get("rows") or []))
            else:
                iterable.append(("Sheet" + str(index), item))
    else:
        raise TypeError("sheets must be a dict or list")

    sheet_count = 0
    row_count = 0
    for raw_name, rows in iterable:
        name = str(raw_name or "Sheet")[:31]
        ws = workbook.create_sheet(title=name)
        sheet_count += 1
        for row in rows or []:
            if isinstance(row, dict):
                values = list(row.values())
            elif isinstance(row, (list, tuple)):
                values = list(row)
            else:
                values = [row]
            ws.append([_cell_value(v) for v in values])
            row_count += 1

    if sheet_count == 0:
        workbook.create_sheet(title="Sheet1")
        sheet_count = 1

    workbook.save(path)
    workbook.close()
    return {"path": path, "sheet_count": sheet_count, "row_count": row_count}
